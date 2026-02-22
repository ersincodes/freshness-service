"""
Document extraction and processing module.

Handles PDF and Excel file ingestion with chunking for RAG retrieval.
Database operations are in repositories/document_repository.py.
Tabular analytics ingestion writes full sheets into SQLite for deterministic queries.
"""
from __future__ import annotations

import hashlib
import logging
import os
import re
import sqlite3
import uuid
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Any

try:
    from pypdf import PdfReader
except ImportError:
    PdfReader = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    import pandas as pd
except ImportError:
    pd = None

logger = logging.getLogger(__name__)


class DocumentType(str, Enum):
    PDF = "pdf"
    XLSX = "xlsx"
    XLS = "xls"


class DocumentStatus(str, Enum):
    PENDING = "pending"
    PROCESSING = "processing"
    READY = "ready"
    ERROR = "error"


@dataclass
class DocumentChunk:
    """A chunk of document content with location metadata."""
    chunk_index: int
    content: str
    metadata: dict[str, Any]


# ============================================================================
# Database Schema
# ============================================================================

CREATE_DOCUMENTS = """
CREATE TABLE IF NOT EXISTS documents (
    document_id TEXT PRIMARY KEY,
    filename TEXT NOT NULL,
    doc_type TEXT NOT NULL,
    size_bytes INTEGER NOT NULL,
    status TEXT NOT NULL DEFAULT 'pending',
    uploaded_at DATETIME NOT NULL,
    error_message TEXT
)
"""

CREATE_DOCUMENT_CHUNKS = """
CREATE TABLE IF NOT EXISTS document_chunks (
    chunk_id TEXT PRIMARY KEY,
    document_id TEXT NOT NULL,
    chunk_index INTEGER NOT NULL,
    content TEXT NOT NULL,
    meta_json TEXT NOT NULL,
    timestamp DATETIME NOT NULL,
    FOREIGN KEY (document_id) REFERENCES documents(document_id) ON DELETE CASCADE
)
"""

CREATE_DOCUMENT_CHUNKS_INDEX = """
CREATE INDEX IF NOT EXISTS idx_document_chunks_document_id 
ON document_chunks(document_id)
"""


def init_document_tables(db_path: str) -> None:
    """Initialize document-related tables in the database."""
    with sqlite3.connect(db_path) as conn:
        cur = conn.cursor()
        cur.execute(CREATE_DOCUMENTS)
        cur.execute(CREATE_DOCUMENT_CHUNKS)
        cur.execute(CREATE_DOCUMENT_CHUNKS_INDEX)
        conn.commit()


# ============================================================================
# Utility Functions
# ============================================================================

def generate_document_id() -> str:
    """Generate a unique document ID."""
    return str(uuid.uuid4())


def hash_chunk_id(document_id: str, chunk_index: int) -> str:
    """Generate a deterministic chunk ID."""
    raw = f"{document_id}:{chunk_index}"
    return hashlib.md5(raw.encode("utf-8")).hexdigest()


def get_document_type_from_filename(filename: str) -> DocumentType | None:
    """Determine document type from filename extension."""
    ext = Path(filename).suffix.lower()
    if ext == ".pdf":
        return DocumentType.PDF
    elif ext == ".xlsx":
        return DocumentType.XLSX
    elif ext == ".xls":
        return DocumentType.XLS
    return None


def validate_mime_type(content_type: str | None, doc_type: DocumentType) -> bool:
    """Validate MIME type matches expected document type."""
    if not content_type:
        return True
    
    valid_types = {
        DocumentType.PDF: ["application/pdf"],
        DocumentType.XLSX: ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"],
        DocumentType.XLS: ["application/vnd.ms-excel", "application/x-msexcel"],
    }
    return content_type in valid_types.get(doc_type, [])


def sanitize_filename(filename: str) -> str:
    """Sanitize filename to prevent path traversal and other issues."""
    filename = os.path.basename(filename)
    filename = re.sub(r'[^\w\-_\. ]', '', filename)
    if len(filename) > 255:
        name, ext = os.path.splitext(filename)
        filename = name[:255 - len(ext)] + ext
    return filename


# ============================================================================
# PDF Extraction
# ============================================================================

def _require_pypdf() -> None:
    if PdfReader is None:
        raise RuntimeError("pypdf is not installed. Install with `pip install pypdf`.")


def extract_pdf_text(file_path: str) -> list[tuple[int, str]]:
    """Extract text from PDF, page by page. Returns list of (page_number, text)."""
    _require_pypdf()
    reader = PdfReader(file_path)
    pages: list[tuple[int, str]] = []
    
    for i, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        text = re.sub(r'\s+', ' ', text).strip()
        if text:
            pages.append((i, text))
    return pages


def chunk_pdf_pages(pages: list[tuple[int, str]], chunk_size: int = 2000) -> list[DocumentChunk]:
    """Chunk PDF pages into smaller pieces with page metadata."""
    chunks: list[DocumentChunk] = []
    chunk_index = 0
    
    for page_num, text in pages:
        if len(text) <= chunk_size:
            chunks.append(DocumentChunk(chunk_index=chunk_index, content=text, metadata={"page": page_num}))
            chunk_index += 1
        else:
            words = text.split()
            current_chunk = []
            current_length = 0
            
            for word in words:
                if current_length + len(word) + 1 > chunk_size and current_chunk:
                    chunks.append(DocumentChunk(
                        chunk_index=chunk_index,
                        content=" ".join(current_chunk),
                        metadata={"page": page_num},
                    ))
                    chunk_index += 1
                    current_chunk = []
                    current_length = 0
                current_chunk.append(word)
                current_length += len(word) + 1
            
            if current_chunk:
                chunks.append(DocumentChunk(
                    chunk_index=chunk_index,
                    content=" ".join(current_chunk),
                    metadata={"page": page_num},
                ))
                chunk_index += 1
    return chunks


# ============================================================================
# Excel Extraction
# ============================================================================

def _require_openpyxl() -> None:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed. Install with `pip install openpyxl`.")


def _require_xlrd() -> None:
    if xlrd is None:
        raise RuntimeError("xlrd is not installed. Install with `pip install xlrd`.")


def extract_xlsx_data(file_path: str) -> list[tuple[str, list[list[Any]]]]:
    """Extract data from XLSX file. Returns list of (sheet_name, rows)."""
    _require_openpyxl()
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheets: list[tuple[str, list[list[Any]]]] = []
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows: list[list[Any]] = []
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append(list(row))
        if rows:
            sheets.append((sheet_name, rows))
    
    workbook.close()
    return sheets


def extract_xls_data(file_path: str) -> list[tuple[str, list[list[Any]]]]:
    """Extract data from XLS file (legacy format). Returns list of (sheet_name, rows)."""
    _require_xlrd()
    workbook = xlrd.open_workbook(file_path)
    sheets: list[tuple[str, list[list[Any]]]] = []
    
    for sheet_idx in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_idx)
        rows: list[list[Any]] = []
        for row_idx in range(sheet.nrows):
            row = sheet.row_values(row_idx)
            if any(cell for cell in row):
                rows.append(row)
        if rows:
            sheets.append((sheet.name, rows))
    return sheets


def _row_to_text(row: list[Any], headers: list[Any] | None = None) -> str:
    """Convert a row to text representation with lossless column mapping.
    
    Handles cases where row has more columns than headers by generating
    synthetic column names for overflow columns.
    """
    if headers:
        parts = []
        header_count = len(headers)
        for i, value in enumerate(row):
            if value is not None and str(value).strip():
                if i < header_count:
                    header_str = str(headers[i]) if headers[i] else f"Col{i+1}"
                else:
                    header_str = f"Col{i+1}"
                parts.append(f"{header_str}={value}")
        return ", ".join(parts)
    return ", ".join(str(v) for v in row if v is not None and str(v).strip())


def chunk_excel_by_budget(
    sheets: list[tuple[str, list[list[Any]]]], char_budget: int = 3500
) -> list[DocumentChunk]:
    """Chunk Excel sheets using character-based budgeting.
    
    Iterates through rows and accumulates them into chunks until the
    character budget is reached, then starts a new chunk. This prevents
    silent row loss that can occur with fixed row counts.
    """
    chunks: list[DocumentChunk] = []
    chunk_index = 0
    
    for sheet_name, rows in sheets:
        if not rows:
            continue
        
        headers = rows[0] if rows else None
        data_rows = rows[1:] if headers else rows
        
        current_lines: list[str] = []
        current_length = 0
        chunk_row_start: int | None = None
        chunk_row_end: int | None = None
        
        for i, row in enumerate(data_rows):
            row_num = i + 2 if headers else i + 1
            row_text = _row_to_text(row, headers)
            if not row_text:
                continue
            
            line = f"Row {row_num}: {row_text}"
            line_len = len(line) + 1
            
            if current_length + line_len > char_budget and current_lines:
                chunks.append(DocumentChunk(
                    chunk_index=chunk_index,
                    content="\n".join(current_lines),
                    metadata={"sheet": sheet_name, "row_start": chunk_row_start, "row_end": chunk_row_end},
                ))
                chunk_index += 1
                current_lines = []
                current_length = 0
                chunk_row_start = None
            
            current_lines.append(line)
            current_length += line_len
            if chunk_row_start is None:
                chunk_row_start = row_num
            chunk_row_end = row_num
        
        if current_lines:
            chunks.append(DocumentChunk(
                chunk_index=chunk_index,
                content="\n".join(current_lines),
                metadata={"sheet": sheet_name, "row_start": chunk_row_start, "row_end": chunk_row_end},
            ))
            chunk_index += 1
    
    return chunks


def chunk_excel_sheets(sheets: list[tuple[str, list[list[Any]]]], rows_per_chunk: int = 50) -> list[DocumentChunk]:
    """Chunk Excel sheets into smaller pieces with sheet/row metadata.
    
    DEPRECATED: Use chunk_excel_by_budget for character-based chunking
    that prevents silent row loss.
    """
    chunks: list[DocumentChunk] = []
    chunk_index = 0
    
    for sheet_name, rows in sheets:
        if not rows:
            continue
        
        headers = rows[0] if rows else None
        data_rows = rows[1:] if headers else rows
        
        for i in range(0, len(data_rows), rows_per_chunk):
            chunk_rows = data_rows[i:i + rows_per_chunk]
            row_start = i + 2 if headers else i + 1
            row_end = row_start + len(chunk_rows) - 1
            
            text_lines = []
            for j, row in enumerate(chunk_rows):
                row_num = row_start + j
                row_text = _row_to_text(row, headers)
                if row_text:
                    text_lines.append(f"Row {row_num}: {row_text}")
            
            if text_lines:
                chunks.append(DocumentChunk(
                    chunk_index=chunk_index,
                    content="\n".join(text_lines),
                    metadata={"sheet": sheet_name, "row_start": row_start, "row_end": row_end},
                ))
                chunk_index += 1
    return chunks


# ============================================================================
# High-Level Processing
# ============================================================================

def process_document(
    file_path: str,
    doc_type: DocumentType,
    chunk_size: int = 2000,
    excel_char_budget: int = 3500,
) -> list[DocumentChunk]:
    """Process a document and return chunks.
    
    Args:
        file_path: Path to the document file.
        doc_type: Type of document (PDF, XLSX, XLS).
        chunk_size: Character budget for PDF chunks.
        excel_char_budget: Character budget for Excel chunks (prevents row loss).
    """
    if doc_type == DocumentType.PDF:
        pages = extract_pdf_text(file_path)
        return chunk_pdf_pages(pages, chunk_size)
    elif doc_type == DocumentType.XLSX:
        sheets = extract_xlsx_data(file_path)
        return chunk_excel_by_budget(sheets, excel_char_budget)
    elif doc_type == DocumentType.XLS:
        sheets = extract_xls_data(file_path)
        return chunk_excel_by_budget(sheets, excel_char_budget)
    else:
        raise ValueError(f"Unsupported document type: {doc_type}")


# ============================================================================
# Tabular Analytics Ingestion (SQLite)
# ============================================================================

def _infer_logical_type(series: Any) -> str:
    """Infer a LogicalType for a pandas Series.

    Priority:
      1. datetime-like dtype OR high parse success → date
      2. boolean-like → boolean
      3. integer-like → integer
      4. float-like → float
      5. fallback → string
    """
    import datetime as _dt

    non_null = series.dropna()
    if non_null.empty:
        return "string"

    # 1. Date detection
    if pd.api.types.is_datetime64_any_dtype(series):
        return "date"
    sample = non_null.iloc[0]
    if isinstance(sample, (_dt.datetime, _dt.date, pd.Timestamp)):
        return "date"
    is_string_dtype = non_null.dtype == object or pd.api.types.is_string_dtype(non_null)

    if is_string_dtype:
        try:
            parsed = pd.to_datetime(non_null, errors="coerce")
            success_ratio = int(parsed.notna().sum()) / len(non_null)
            if success_ratio >= 0.8:
                return "date"
        except Exception:
            pass

    # 2. Boolean detection
    _BOOL_VALS = {"true", "false", "yes", "no", "0", "1"}
    if non_null.dtype == bool or (
        is_string_dtype
        and all(str(v).strip().lower() in _BOOL_VALS for v in non_null)
    ):
        return "boolean"

    # 3-4. Numeric detection
    if pd.api.types.is_integer_dtype(series):
        return "integer"
    if pd.api.types.is_float_dtype(series):
        if (non_null == non_null.astype(int)).all():
            return "integer"
        return "float"
    if is_string_dtype:
        coerced = pd.to_numeric(non_null, errors="coerce")
        if coerced.notna().sum() / len(non_null) >= 0.9:
            if (coerced.dropna() == coerced.dropna().astype(int)).all():
                return "integer"
            return "float"

    return "string"


def _normalize_cell_value(x: Any, logical_type: str) -> Any:
    """Normalize a cell value according to its logical type.

    - date → epoch seconds (int, UTC)
    - boolean → 0/1 (int)
    - integer → int
    - float → float
    - string → str (trimmed)
    """
    if x is None:
        return None
    if pd is not None and pd.isna(x):
        return None

    import datetime as _dt
    from datetime import timezone as _tz

    if logical_type == "date":
        if isinstance(x, pd.Timestamp):
            if x.tzinfo is None:
                x = x.tz_localize("UTC")
            return int(x.timestamp())
        if isinstance(x, _dt.datetime):
            if x.tzinfo is None:
                x = x.replace(tzinfo=_tz.utc)
            return int(x.timestamp())
        if isinstance(x, _dt.date):
            return int(_dt.datetime(x.year, x.month, x.day, tzinfo=_tz.utc).timestamp())
        try:
            parsed = pd.to_datetime(x)
            if parsed.tzinfo is None:
                parsed = parsed.tz_localize("UTC")
            return int(parsed.timestamp())
        except Exception:
            return None

    if logical_type == "boolean":
        s = str(x).strip().lower()
        return 1 if s in {"true", "yes", "1", "1.0"} else 0

    if logical_type == "integer":
        try:
            return int(float(x))
        except (ValueError, TypeError):
            return None

    if logical_type == "float":
        try:
            return float(x)
        except (ValueError, TypeError):
            return None

    # string
    return str(x).strip()


def ingest_excel_to_sqlite(
    *,
    excel_path: str,
    document_id: str,
    sqlite_connection: sqlite3.Connection,
) -> None:
    """Ingest all Excel sheets into typed SQLite tables with profiling.

    Dates are stored as epoch-second INTEGERs.  Booleans as 0/1 INTEGERs.
    Numeric columns keep their native type.  Strings remain TEXT.
    Column metadata and dataset profiles are persisted for downstream
    validation and deterministic query compilation.
    """
    if pd is None:
        logger.warning("pandas is not installed — tabular analytics ingestion skipped")
        return

    from .analytics.metadata_repository import MetadataRepository
    from .analytics.models import ColumnMetadata, SQLITE_TYPE_MAP
    from .analytics.profiler import profile_dataframe

    meta_repo = MetadataRepository(sqlite_connection)

    sheets: dict[str, Any] = pd.read_excel(excel_path, sheet_name=None)
    if not sheets:
        raise ValueError("No sheets found in workbook")

    default_sheet_name = next(iter(sheets.keys()))

    for sheet_name, df in sheets.items():
        if df is None or df.empty:
            continue

        original_headers = [str(c) for c in df.columns]
        augmented_headers = ["_source_row_number", *original_headers]

        # Infer logical types on the raw DataFrame
        col_logical_types: dict[str, str] = {"_source_row_number": "integer"}
        for header in original_headers:
            col_logical_types[header] = _infer_logical_type(df[header])

        logger.info(
            "Sheet '%s' column types: %s",
            sheet_name,
            {h: col_logical_types[h] for h in original_headers},
        )

        df2 = df.copy()
        df2.insert(0, "_source_row_number", range(1, len(df2) + 1))

        table_name = _build_document_sheet_table_name(
            document_id=document_id, sheet_name=str(sheet_name)
        )
        original_to_safe = _build_safe_column_mapping(augmented_headers)

        df2.columns = [original_to_safe[h] for h in augmented_headers]
        df2 = df2.astype(object).where(pd.notnull(df2), None)

        # Typed normalization
        for header in augmented_headers:
            safe = original_to_safe[header]
            ltype = col_logical_types[header]
            df2[safe] = df2[safe].map(lambda x, lt=ltype: _normalize_cell_value(x, lt))

        # Build column metadata
        col_meta_list: list[ColumnMetadata] = []
        for h in augmented_headers:
            ltype = col_logical_types[h]
            sqlite_type = SQLITE_TYPE_MAP.get(ltype, "TEXT")
            nullable = h != "_source_row_number"
            col_meta_list.append(ColumnMetadata(
                column_name=h,
                logical_type=ltype,
                sqlite_type=sqlite_type,
                nullable=nullable,
                original_name=h,
                safe_name=original_to_safe[h],
            ))

        col_meta_dict = {m.original_name: m for m in col_meta_list}

        # Create typed table + indices
        _drop_and_create_typed_table(
            sqlite_connection=sqlite_connection,
            table_name=table_name,
            columns=col_meta_list,
        )

        safe_cols = [original_to_safe[h] for h in augmented_headers]
        _bulk_insert(
            sqlite_connection=sqlite_connection,
            table_name=table_name,
            safe_columns=safe_cols,
            rows=df2.itertuples(index=False, name=None),
        )

        # Register metadata
        meta_repo.register_table(document_id, str(sheet_name), table_name, len(df2))
        meta_repo.register_columns(document_id, str(sheet_name), col_meta_list)
        if str(sheet_name) == str(default_sheet_name):
            meta_repo.register_default_sheet(document_id, str(sheet_name))

        # Compute and persist profile
        try:
            profile = profile_dataframe(df2, col_meta_dict)
            meta_repo.upsert_profile(document_id, str(sheet_name), profile)
        except Exception as exc:
            logger.warning("Profiling failed for sheet '%s': %s", sheet_name, exc)

    logger.info("Ingested %d sheet(s) for document %s into SQLite", len(sheets), document_id)


def _build_document_sheet_table_name(*, document_id: str, sheet_name: str) -> str:
    doc_part = re.sub(r"[^a-zA-Z0-9_]+", "_", document_id)[:24].strip("_") or "doc"
    sheet_hash = hashlib.sha1(sheet_name.encode("utf-8")).hexdigest()[:10]
    return f"doc_{doc_part}__{sheet_hash}"


def _build_safe_column_mapping(original_headers: list[str]) -> dict[str, str]:
    used: set[str] = set()
    mapping: dict[str, str] = {}

    for raw in original_headers:
        base = re.sub(r"[^a-zA-Z0-9_]+", "_", str(raw).strip().lower())
        base = re.sub(r"_+", "_", base).strip("_")
        base = base or "col"

        candidate = f"col_{base}"
        if candidate[0].isdigit():
            candidate = f"col_{candidate}"

        unique = candidate
        suffix = 2
        while unique in used:
            unique = f"{candidate}_{suffix}"
            suffix += 1

        used.add(unique)
        mapping[str(raw)] = unique

    return mapping


def _drop_and_create_typed_table(
    *,
    sqlite_connection: sqlite3.Connection,
    table_name: str,
    columns: list,
) -> None:
    """Create a table with explicit SQLite types and useful indices."""
    from .analytics.models import ColumnMetadata  # noqa: used for type hint

    columns_ddl = ", ".join(f"{c.safe_name} {c.sqlite_type}" for c in columns)

    with sqlite_connection:
        sqlite_connection.execute(f"DROP TABLE IF EXISTS {table_name};")
        sqlite_connection.execute(f"CREATE TABLE {table_name} ({columns_ddl});")

        for col in columns:
            if "source_row_number" in col.safe_name:
                sqlite_connection.execute(
                    f"CREATE INDEX IF NOT EXISTS idx_{table_name}__rownum "
                    f"ON {table_name} ({col.safe_name});"
                )
            elif col.logical_type == "date":
                sqlite_connection.execute(
                    f"CREATE INDEX IF NOT EXISTS idx_{table_name}__{col.safe_name} "
                    f"ON {table_name} ({col.safe_name});"
                )
            elif any(kw in col.original_name.lower() for kw in ("_id", "id", "code", "index")):
                sqlite_connection.execute(
                    f"CREATE INDEX IF NOT EXISTS idx_{table_name}__{col.safe_name} "
                    f"ON {table_name} ({col.safe_name});"
                )


def _bulk_insert(
    *,
    sqlite_connection: sqlite3.Connection,
    table_name: str,
    safe_columns: list[str],
    rows: Any,
) -> None:
    placeholders = ",".join(["?"] * len(safe_columns))
    cols_sql = ",".join(safe_columns)
    sql = f"INSERT INTO {table_name} ({cols_sql}) VALUES ({placeholders});"

    with sqlite_connection:
        sqlite_connection.executemany(sql, list(rows))
